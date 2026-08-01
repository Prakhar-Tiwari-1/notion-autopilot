import os
import re
import json
import html
import base64
import random
import smtplib
import requests
from datetime import date, timedelta
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from openai import OpenAI
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

QUOTES_FILE       = os.path.join(os.path.dirname(__file__), "quotes.xlsx")
QUOTE_STATE_FILE  = os.path.join(os.path.dirname(__file__), "quote_state.json")

NOTION_API_KEY   = os.environ["NOTION_API_KEY"]
OPENAI_API_KEY   = os.environ["OPENAI_API_KEY"]
ZOHO_PASSWORD    = os.environ["ZOHO_API_KEY"]
TASK_DATABASE_ID = "911d2703b4ea4135ac8d75f17cacd3ba"

FROM_EMAIL = "contact@prakhartiwari.com"
TO_EMAIL   = "prakhar.tiwari@polytechnique.edu"

NOTION_HEADERS = {
    "Authorization": f"Bearer {NOTION_API_KEY}",
    "Notion-Version": "2022-06-28",
    "Content-Type": "application/json",
}


def fetch_comments(page_id: str) -> str:
    res = requests.get(
        "https://api.notion.com/v1/comments",
        headers=NOTION_HEADERS,
        params={"block_id": page_id},
    )
    if not res.ok:
        return ""
    comments = []
    for c in res.json().get("results", []):
        text = "".join(rt["plain_text"] for rt in c.get("rich_text", []))
        if text.strip():
            comments.append(f"    💬 {text}")
    return "\n".join(comments)


def fetch_tasks() -> str:
    url      = f"https://api.notion.com/v1/databases/{TASK_DATABASE_ID}/query"
    in_14    = (date.today() + timedelta(days=14)).isoformat()
    tasks    = []

    payload = {
        "sorts": [{"property": "Due Date", "direction": "ascending"}],
        "filter": {
            "and": [
                {"property": "Stage", "select": {"does_not_equal": "✅ Finished"}},
                {
                    "or": [
                        {"property": "Priority", "select": {"equals": "🔥 Critical"}},
                        {"property": "Priority", "select": {"equals": "⚡ High"}},
                        {"property": "Due Date", "date": {"on_or_before": in_14}},
                    ]
                },
            ]
        },
    }

    while True:
        res = requests.post(url, headers=NOTION_HEADERS, json=payload)
        res.raise_for_status()
        data = res.json()

        for page in data.get("results", []):
            props    = page["properties"]
            page_id  = page["id"]

            name     = "".join(t["plain_text"] for t in props["Task"]["title"])
            priority = (props["Priority"]["select"] or {}).get("name", "—")
            stage    = (props["Stage"]["select"] or {}).get("name", "—")
            category = (props["Category"]["select"] or {}).get("name", "—")
            due      = (props["Due Date"]["date"] or {}).get("start", "no date")
            notes    = "".join(t["plain_text"] for t in props["Notes"]["rich_text"])
            comments = fetch_comments(page_id)

            line = f"[{priority}] {name}\n  Stage: {stage} | Category: {category} | Due: {due}"
            if notes:
                line += f"\n  Notes: {notes}"
            if comments:
                line += f"\n{comments}"
            tasks.append(line)

        if not data.get("has_more"):
            break
        payload["start_cursor"] = data["next_cursor"]

    return "\n\n".join(tasks)


def next_quote() -> dict:
    """Pull the next quote from quotes.xlsx in a shuffled, non-repeating order.

    quote_state.json tracks the shuffled order and current position so each
    of the 500 quotes is used exactly once before any repeat. When the order
    is exhausted (or missing), it's reshuffled and position resets to 0.
    """
    wb = load_workbook(QUOTES_FILE, read_only=True)
    rows = list(wb.active.iter_rows(min_row=2, values_only=True))
    quotes = [{"text": text, "author": author} for _id, text, author in rows]

    try:
        with open(QUOTE_STATE_FILE, "r", encoding="utf-8") as f:
            state = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        state = {"order": [], "position": 0}

    order, position = state.get("order", []), state.get("position", 0)
    if position >= len(order):
        order = list(range(len(quotes)))
        random.shuffle(order)
        position = 0

    quote = quotes[order[position]]

    with open(QUOTE_STATE_FILE, "w", encoding="utf-8") as f:
        json.dump({"order": order, "position": position + 1}, f)

    return quote


def build_content(tasks: str) -> dict:
    today  = date.today().strftime("%A, %B %d %Y")
    client = OpenAI(api_key=OPENAI_API_KEY)

    template = f"""\
You are a sharp personal assistant. Organise the user's raw task list into a JSON object.
You decide nothing about presentation — only the content. Today is {today}.

Return JSON with EXACTLY this shape:

{{
  "critical": [
    {{
      "name": "Task name",
      "due": "Due date as written, or empty string",
      "category": "Category, or empty string",
      "notes": "Notes if any, else empty string",
      "comment": "Comment if any, else empty string"
    }}
  ],
  "upcoming": [ {{ "name": "", "due": "", "category": "", "notes": "", "comment": "" }} ],
  "high_no_deadline": [ {{ "name": "", "due": "", "category": "", "notes": "", "comment": "" }} ],
  "weekly_plan": [
    {{ "days": "Mon – Wed", "focus": "Direct, specific focus for these days" }},
    {{ "days": "Thu – Fri", "focus": "..." }},
    {{ "days": "Weekend",   "focus": "Anything to wrap up or prep" }}
  ],
  "words": {{
    "excerpt": "A short excerpt (3-6 sentences) from a real speech or writing by a well-known world leader, entrepreneur, athlete, philosopher, or artist. Pick a different person every single time, varying across fields and eras. Never the same person twice in a row.",
    "author": "Full name",
    "context": "e.g. Stanford Commencement 2005 / Letter to shareholders 1997"
  }}
}}

Sorting rules:
- "critical"          = tasks marked Critical or otherwise urgent.
- "upcoming"          = everything else due within the next two weeks.
- "high_no_deadline"  = High priority tasks with no due date.
- Put every task in exactly one bucket. Drop nothing — every note and comment must travel with its task.
- Return an empty array for any bucket with no tasks.
- The weekly plan must be practical and specific to these tasks — no generic advice, 3-5 entries.
- Output JSON only.
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": template},
            {"role": "user",   "content": tasks},
        ],
    )
    return json.loads(response.choices[0].message.content)


# Warm editorial palette — beige, brown, cream
PAGE_BG     = "#e7dcc8"
PANEL       = "#fbf6ec"
CARD        = "#f4ebd9"
BORDER      = "#ddcdb0"
ACCENT      = "#9b5e34"
ACCENT_SOFT = "#b07d4a"
TEXT        = "#3c2f23"
MUTED       = "#6e5d47"
FAINT       = "#9a886e"

DISPLAY = "'Open Sans', 'Helvetica Neue', Helvetica, Arial, sans-serif"
BODY    = "'Open Sans', 'Helvetica Neue', Helvetica, Arial, sans-serif"


# json.loads sometimes yields literal "🎓" text instead of the emoji
# when the model double-escapes in its JSON output.
_ESC_SURROGATE = re.compile(r"\\u(d[89ab][0-9a-f]{2})\\u(d[c-f][0-9a-f]{2})", re.I)
_ESC_ASTRAL    = re.compile(r"\\U([0-9a-fA-F]{8})")
_ESC_BMP       = re.compile(r"\\u([0-9a-fA-F]{4})")


def _decode_escapes(text: str) -> str:
    if not text or "\\" not in text:
        return text

    def _pair(m):
        hi, lo = int(m.group(1), 16), int(m.group(2), 16)
        return chr(0x10000 + ((hi - 0xD800) << 10) + (lo - 0xDC00))

    text = _ESC_SURROGATE.sub(_pair, text)
    text = _ESC_ASTRAL.sub(lambda m: chr(int(m.group(1), 16)), text)
    text = _ESC_BMP.sub(lambda m: chr(int(m.group(1), 16)), text)
    return text


def _ascii_safe(markup: str) -> str:
    # Emoji as raw 4-byte UTF-8 get mangled by some mail relays.
    # Numeric HTML entities render reliably everywhere.
    return "".join(c if ord(c) < 128 else f"&#x{ord(c):X};" for c in markup)


def _esc(text: str) -> str:
    return html.escape(_decode_escapes(text or "")).strip()


# Webmail won't render font emoji, so we embed Twemoji PNGs as CID attachments.
TWEMOJI_BASE = "https://cdn.jsdelivr.net/gh/jdecked/twemoji@15.1.0/assets/72x72/"

# Excludes Geometric Shapes (◈ ◇) which are used as plain-text bullets.
EMOJI_RE = re.compile(
    "(?:[\U0001F300-\U0001FAFF\U0001F000-\U0001F0FF\U0001F1E6-\U0001F1FF"
    "\U00002600-\U000027BF\U00002300-\U000023FF\U00002B00-\U00002BFF"
    "\U0000FE0F\U0000200D\U0001F3FB-\U0001F3FF]+)"
)

_emoji_cache: dict[str, bytes | None] = {}


def _twemoji_name(cluster: str) -> str | None:
    """Twemoji filename stem: hyphen-joined hex codepoints, sans FE0F."""
    cps = [ord(c) for c in cluster if ord(c) != 0xFE0F]
    return "-".join(f"{c:x}" for c in cps) if cps else None


def _fetch_emoji_png(name: str) -> bytes | None:
    if name not in _emoji_cache:
        try:
            r = requests.get(f"{TWEMOJI_BASE}{name}.png", timeout=10)
            _emoji_cache[name] = r.content if r.ok else None
        except requests.RequestException:
            _emoji_cache[name] = None
    return _emoji_cache[name]


def inline_emoji(markup: str, mode: str = "cid") -> tuple[str, dict]:
    """Replace emoji characters with Twemoji <img> tags.

    mode="cid"  → cid: references; caller must attach the returned PNGs.
    mode="data" → base64 data URIs, suitable for local browser previews.
    """
    images: dict[str, bytes] = {}

    def repl(match: "re.Match") -> str:
        cluster = match.group(0)
        name = _twemoji_name(cluster)
        if not name:
            return cluster
        png = _fetch_emoji_png(name)
        if png is None:
            return cluster
        if mode == "data":
            src = "data:image/png;base64," + base64.b64encode(png).decode()
        else:
            cid = f"emoji-{name}"
            images[cid] = png
            src = f"cid:{cid}"
        return (f'<img src="{src}" alt="" width="18" height="18" '
                'style="height:1.05em;width:auto;vertical-align:-0.18em;'
                'margin:0 2px;border:0;">')

    return EMOJI_RE.sub(repl, markup), images


def _task_card(task: dict, marker: str, accent: str) -> str:
    name     = _esc(task.get("name"))
    if not name:
        return ""
    due      = _esc(task.get("due"))
    category = _esc(task.get("category"))
    notes    = _esc(task.get("notes"))
    comment  = _esc(task.get("comment"))

    meta = []
    if due:
        meta.append(f"<span style='color:{ACCENT};'>Due</span> "
                    f"<span style='color:{FAINT};'>·</span> {due}")
    if category:
        meta.append(f"<span style='color:{ACCENT};'>Category</span> "
                    f"<span style='color:{FAINT};'>·</span> {category}")
    meta_html = (
        f"<div style='font-family:{BODY};font-size:13px;color:{MUTED};"
        f"margin-top:6px;letter-spacing:.2px;'>{'&nbsp;&nbsp;&nbsp;'.join(meta)}</div>"
        if meta else ""
    )
    notes_html = (
        f"<div style='font-family:{BODY};font-size:13.5px;color:{MUTED};"
        f"font-style:italic;margin-top:8px;line-height:1.5;'>{notes}</div>"
        if notes else ""
    )
    comment_html = (
        f"<div style='font-family:{BODY};font-size:13px;color:{ACCENT_SOFT};"
        f"margin-top:8px;line-height:1.5;'>💬&nbsp; {comment}</div>"
        if comment else ""
    )

    return f"""
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin:0 0 12px 0;">
      <tr>
        <td style="background:{CARD};border:1px solid {BORDER};border-left:3px solid {accent};
                   border-radius:8px;padding:16px 20px;">
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td valign="top" width="26" style="font-family:{DISPLAY};font-size:19px;
                         line-height:1.3;color:{accent};">{marker}</td>
              <td valign="top">
                <div style="font-family:{DISPLAY};font-size:19px;font-weight:600;color:{TEXT};
                            line-height:1.3;letter-spacing:.2px;">{name}</div>
                {meta_html}
                {notes_html}
                {comment_html}
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>"""


def _section(label: str, emoji: str, tasks: list, marker: str, accent: str) -> str:
    cards = "".join(_task_card(t, marker, accent) for t in (tasks or []))
    if not cards.strip():
        return ""
    return f"""
    <tr><td style="padding:0 32px;">
      <div style="font-family:{BODY};font-size:12px;font-weight:700;letter-spacing:3px;
                  text-transform:uppercase;color:{ACCENT};padding:8px 0 4px 0;">
        {emoji}&nbsp;&nbsp;{label}
      </div>
      <div style="height:1px;background:linear-gradient(90deg,{ACCENT}44,{BORDER}00);
                  margin:0 0 16px 0;"></div>
      {cards}
    </td></tr>
    <tr><td style="height:14px;line-height:14px;font-size:0;">&nbsp;</td></tr>"""


def render_html(data: dict, emoji_mode: str = "cid") -> tuple[str, dict]:
    today      = date.today().strftime("%A, %B %d %Y")
    week_start = date.today().strftime("%B %d")
    week_end   = (date.today() + timedelta(days=6)).strftime("%B %d")

    quote = data.get("quote") or {}
    words = data.get("words") or {}

    quote_html = ""
    if _esc(quote.get("text")):
        quote_html = f"""
        <tr><td style="padding:0 32px 8px 32px;">
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
            <tr><td style="background:{CARD};border:1px solid {BORDER};border-radius:8px;
                           padding:22px 26px;">
              <div style="font-family:{BODY};font-size:11px;font-weight:700;letter-spacing:3px;
                          text-transform:uppercase;color:{ACCENT};margin-bottom:10px;">
                ✦&nbsp;&nbsp;Quote of the Day
              </div>
              <div style="font-family:{DISPLAY};font-size:22px;font-style:italic;color:{TEXT};
                          line-height:1.45;">“{_esc(quote.get('text'))}”</div>
              <div style="font-family:{BODY};font-size:13px;color:{MUTED};margin-top:12px;
                          letter-spacing:.5px;">— {_esc(quote.get('author'))}</div>
            </td></tr>
          </table>
        </td></tr>
        <tr><td style="height:18px;line-height:18px;font-size:0;">&nbsp;</td></tr>"""

    sections  = ""
    sections += _section("Critical & Urgent",        "🔥", data.get("critical"),         "◈", ACCENT)
    sections += _section("Coming Up — Next 2 Weeks",  "📅", data.get("upcoming"),         "◇", "#bfa06f")
    sections += _section("High Priority · No Deadline", "⚡", data.get("high_no_deadline"), "◇", "#9c7b4f")

    plan_rows = ""
    for item in (data.get("weekly_plan") or []):
        days  = _esc(item.get("days"))
        focus = _esc(item.get("focus"))
        if not (days or focus):
            continue
        plan_rows += f"""
          <tr>
            <td style="font-family:{BODY};font-size:13px;font-weight:700;color:{ACCENT};
                       letter-spacing:1px;padding:9px 16px 9px 0;white-space:nowrap;
                       vertical-align:top;border-bottom:1px solid {BORDER};">{days}</td>
            <td style="font-family:{BODY};font-size:14px;color:{TEXT};line-height:1.5;
                       padding:9px 0;border-bottom:1px solid {BORDER};">{focus}</td>
          </tr>"""
    plan_html = ""
    if plan_rows:
        plan_html = f"""
        <tr><td style="padding:0 32px;">
          <div style="font-family:{BODY};font-size:12px;font-weight:700;letter-spacing:3px;
                      text-transform:uppercase;color:{ACCENT};padding:8px 0 4px 0;">
            🗓&nbsp;&nbsp;Weekly Plan&nbsp;&nbsp;·&nbsp;&nbsp;{week_start} – {week_end}
          </div>
          <div style="height:1px;background:linear-gradient(90deg,{ACCENT}44,{BORDER}00);
                      margin:0 0 12px 0;"></div>
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0">{plan_rows}
          </table>
        </td></tr>
        <tr><td style="height:26px;line-height:26px;font-size:0;">&nbsp;</td></tr>"""

    words_html = ""
    if _esc(words.get("excerpt")):
        attribution = _esc(words.get("author"))
        context     = _esc(words.get("context"))
        if context:
            attribution += f", <span style='color:{FAINT};'>{context}</span>"
        words_html = f"""
        <tr><td style="padding:0 32px;">
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
            <tr><td style="background:linear-gradient(135deg,#f3e6cf,#efe1c6);
                           border:1px solid {BORDER};border-radius:8px;padding:26px 28px;">
              <div style="font-family:{BODY};font-size:11px;font-weight:700;letter-spacing:3px;
                          text-transform:uppercase;color:{ACCENT};margin-bottom:14px;">
                🎙&nbsp;&nbsp;Words to Carry
              </div>
              <div style="font-family:{DISPLAY};font-size:20px;font-style:italic;color:{TEXT};
                          line-height:1.55;">“{_esc(words.get('excerpt'))}”</div>
              <div style="font-family:{BODY};font-size:13px;color:{MUTED};margin-top:14px;
                          letter-spacing:.5px;">— {attribution}</div>
            </td></tr>
          </table>
        </td></tr>
        <tr><td style="height:26px;line-height:26px;font-size:0;">&nbsp;</td></tr>"""

    markup = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta name="color-scheme" content="light">
  <meta name="supported-color-schemes" content="light">
  <link href="https://fonts.googleapis.com/css2?family=Open+Sans:ital,wght@0,400;0,600;0,700;1,400&display=swap" rel="stylesheet">
  <title>Mind Tracker</title>
</head>
<body style="margin:0;padding:0;background:{PAGE_BG};">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
         style="background:{PAGE_BG};margin:0;padding:0;">
    <tr><td align="center" style="padding:28px 14px;">

      <table role="presentation" width="600" cellpadding="0" cellspacing="0"
             style="width:600px;max-width:600px;background:{PANEL};border:1px solid {BORDER};
                    border-radius:14px;overflow:hidden;">

        <!-- Header -->
        <tr><td style="padding:38px 32px 8px 32px;text-align:center;">
          <div style="font-family:{DISPLAY};font-size:13px;letter-spacing:6px;color:{ACCENT};
                      text-transform:uppercase;margin-bottom:10px;">🧠 &nbsp;Mind Tracker</div>
          <div style="font-family:{DISPLAY};font-size:34px;font-weight:600;color:{TEXT};
                      letter-spacing:1px;line-height:1.15;">DS&rsquo;s Daily Brief</div>
          <div style="font-family:{BODY};font-size:13px;color:{MUTED};margin-top:10px;
                      letter-spacing:2px;text-transform:uppercase;">{today}</div>
          <div style="width:54px;height:2px;background:{ACCENT};margin:20px auto 0 auto;"></div>
        </td></tr>
        <tr><td style="height:26px;line-height:26px;font-size:0;">&nbsp;</td></tr>

        {quote_html}
        {sections}
        {plan_html}
        {words_html}

        <!-- Footer -->
        <tr><td style="background:{CARD};border-top:1px solid {BORDER};padding:24px 32px;
                       text-align:center;">
          <div style="font-family:{DISPLAY};font-size:17px;color:{ACCENT_SOFT};letter-spacing:1px;
                      font-style:italic;">Stay sharp.</div>
        </td></tr>

      </table>

      <div style="font-family:{BODY};font-size:11px;color:{MUTED};margin-top:18px;
                  letter-spacing:1px;">Generated from Notion · {today}</div>

    </td></tr>
  </table>
</body>
</html>"""
    markup, images = inline_emoji(markup, emoji_mode)
    return _ascii_safe(markup), images


def render_text(data: dict) -> str:
    today      = date.today().strftime("%A, %B %d %Y")
    week_start = date.today().strftime("%B %d")
    week_end   = (date.today() + timedelta(days=6)).strftime("%B %d")
    lines      = []

    def clean(value) -> str:
        return _decode_escapes((value or "").strip())

    lines.append("🧠  DS'S MIND TRACKER")
    lines.append(today)
    lines.append("=" * 38)

    quote = data.get("quote") or {}
    if quote.get("text"):
        lines += ["", "✦  QUOTE OF THE DAY", "",
                  f'  "{clean(quote.get("text"))}"',
                  f'  — {clean(quote.get("author"))}']

    def block(label, tasks, marker):
        nonlocal lines
        if not tasks:
            return
        lines += ["", "=" * 38, "", f"  {label}", "  " + "-" * len(label)]
        for t in tasks:
            name = clean(t.get("name"))
            if not name:
                continue
            lines.append(f"  {marker}  {name}")
            meta = []
            if t.get("due"):
                meta.append(f"Due · {clean(t.get('due'))}")
            if t.get("category"):
                meta.append(f"Category · {clean(t.get('category'))}")
            if meta:
                lines.append("     " + "   ".join(meta))
            if t.get("notes"):
                lines.append(f"     {clean(t.get('notes'))}")
            if t.get("comment"):
                lines.append(f"     💬 {clean(t.get('comment'))}")
            lines.append("")

    block("🔥  CRITICAL & URGENT",            data.get("critical"),         "◈")
    block("📅  COMING UP — NEXT 2 WEEKS",      data.get("upcoming"),         "◇")
    block("⚡  HIGH PRIORITY · NO DEADLINE",   data.get("high_no_deadline"), "◇")

    plan = data.get("weekly_plan") or []
    if plan:
        lines += ["", "=" * 38, "", f"  🗓  WEEKLY PLAN · {week_start} – {week_end}",
                  "  " + "-" * 26]
        for item in plan:
            days  = clean(item.get("days"))
            focus = clean(item.get("focus"))
            lines.append(f"  {days:<12} · {focus}")

    words = data.get("words") or {}
    if words.get("excerpt"):
        attribution = clean(words.get("author"))
        if words.get("context"):
            attribution += f", {clean(words.get('context'))}"
        lines += ["", "=" * 38, "", "  🎙  WORDS TO CARRY", "",
                  f'  "{clean(words.get("excerpt"))}"', "",
                  f"  — {attribution}"]

    lines += ["", "=" * 38, "  Stay sharp. 🚀", "=" * 38]
    return "\n".join(lines)


def send_email(text_body: str, html_body: str, images: dict | None = None) -> None:
    today = date.today().strftime("%b %d")
    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"🧠 Mind Tracker — {today}"
    msg["From"]    = FROM_EMAIL
    msg["To"]      = TO_EMAIL

    # Order matters: the last part is the client's preferred rendering.
    msg.attach(MIMEText(text_body, "plain", "utf-8"))

    # HTML + its inline emoji images go in a multipart/related so the
    # cid: references resolve against the attached PNGs.
    related = MIMEMultipart("related")
    related.attach(MIMEText(html_body, "html", "utf-8"))
    for cid, png in (images or {}).items():
        img = MIMEImage(png, _subtype="png")
        img.add_header("Content-ID", f"<{cid}>")
        img.add_header("Content-Disposition", "inline", filename=f"{cid}.png")
        related.attach(img)
    msg.attach(related)

    with smtplib.SMTP_SSL("smtp.zoho.eu", 465) as server:
        server.login(FROM_EMAIL, ZOHO_PASSWORD)
        server.sendmail(FROM_EMAIL, TO_EMAIL, msg.as_string())

    print(f"Email sent to {TO_EMAIL}")


if __name__ == "__main__":
    print("Fetching tasks from Notion...")
    tasks = fetch_tasks()

    if not tasks.strip():
        print("No urgent or upcoming tasks found.")
        raise SystemExit(0)

    print("Building email...")
    data = build_content(tasks)
    data["quote"] = next_quote()
    html_body, emoji_images = render_html(data)
    text_body = render_text(data)

    print("Sending email...")
    send_email(text_body, html_body, emoji_images)

    print("Done.")
